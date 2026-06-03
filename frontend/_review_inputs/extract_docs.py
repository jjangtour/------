import json
import sys
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader


def extract_pdf(path: Path):
    reader = PdfReader(str(path))
    pages = []
    for i, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        pages.append({"page": i, "text": "\n".join(line.strip() for line in text.splitlines() if line.strip())})
    return pages


def extract_xlsx(path: Path):
    wb = load_workbook(path, data_only=False)
    result = {}
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(values_only=True):
            values = ["" if value is None else str(value) for value in row]
            while values and values[-1] == "":
                values.pop()
            rows.append(values)
        while rows and not rows[-1]:
            rows.pop()
        result[ws.title] = rows
    return result


def main():
    base = Path(sys.argv[1])
    payload = {
        "pdf": extract_pdf(base / "proposal.pdf"),
        "xlsx": extract_xlsx(base / "feature_def.xlsx"),
    }
    out = base / "extracted.json"
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(out)


if __name__ == "__main__":
    main()
